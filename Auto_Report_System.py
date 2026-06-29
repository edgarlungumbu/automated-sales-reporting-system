# %%
import sys
from pathlib import Path
import pandas as pd
from datetime import date

input_file =Path(sys.argv[1]) if len(sys.argv)>1 else Path('raw_sales_data.xlsx')
output_file = Path(sys.argv[2]) if len(sys.argv)>2 else Path(f'sales_report_{date.today():%Y-%m-%d}.xlsx')

# %%
#Data Clean
import numpy as np
def clean_data(df):
    df = df.drop_duplicates()


    df['Product'] =df['Product'].fillna('Unknown')
    df['Category'] = df['Category'].fillna('Unknown')
    df['Clean Category'] = df['Category'].str.title()

    df['Clean Date'] = pd.to_datetime(df['Date'],dayfirst=True,errors='coerce')
    #Changes to consistent datetime format
    df['Month'] = df['Clean Date'].dt.strftime('%B')
    df['Day_of_Week'] =  df['Clean Date'].dt.day_name()


    df['Unit_Price'] =  df['Unit_Price'].replace([np.inf, - np.inf, np.nan],0)
    df['Quantity'] =  df['Quantity'].replace([np.inf, - np.inf, np.nan],0)
    df['Revenue'] =  df['Revenue'].replace([np.inf, - np.inf, np.nan],0)


    df['Quantity'] = df['Quantity'].astype(int)
    df['Unit_Price'] = df['Unit_Price'].astype(float)
    df['Revenue'] = df['Revenue'].astype(float)
    df['Revenue'] = df['Quantity'] * df['Unit_Price']

    return df




# %%
#Summary Metrics
def build_summaries(df):
    Category_Summary =df.groupby('Clean Category').agg(
        Order_Count = ('Order_ID', 'count'),
        Total_Revenue = ('Revenue', 'sum'),

    )

    top_products =(df.groupby('Product')['Revenue']
        .sum()
        .sort_values(ascending=False)
        .reset_index()
                     )

    top_product = top_products.iloc[0]['Product']

    Revenue_by_Category = (df.groupby('Clean Category')['Revenue']
        .sum()
        .sort_values(ascending=False)
        .reset_index()
                           )

    Revenue_by_Month = (df.groupby('Month')['Revenue']
        .sum()
        .reset_index()
                           )
    month_order = [
    'January','February','March','April','May','June',
    'July','August','September','October','November','December'
]

    Revenue_by_Month['Month'] = pd.Categorical(
        Revenue_by_Month['Month'],
        categories=month_order,
        ordered=True
    )

    Revenue_by_Month = Revenue_by_Month.sort_values('Month')
    Revenue_by_Day = (df.groupby('Day_of_Week')['Revenue']
        .sum()
        .reset_index()
                       )
    day_order = [
        'Monday',
        'Tuesday',
        'Wednesday',
        'Thursday',
        'Friday',
        'Saturday',
        'Sunday'
    ]
    Revenue_by_Day['Day_of_Week'] = pd.Categorical(
        Revenue_by_Day['Day_of_Week'],
        categories=day_order,
        ordered=True
    )
    Revenue_by_Day = Revenue_by_Day.sort_values('Day_of_Week')

    return Category_Summary,Revenue_by_Category,Revenue_by_Month,Revenue_by_Day, top_products,top_product



# %%
#Overall Summaries
def print_summary(df,Revenue_by_Month, Revenue_by_Category, top_product):
    top_products =(df.groupby('Product')['Revenue']
        .sum()
        .sort_values(ascending=False)
        .reset_index()
                     )
    total_revenue =  df['Revenue'].sum()
    total_orders = len(df)
    avg_order_value = total_revenue/total_orders
    best_month = Revenue_by_Month.loc[Revenue_by_Month['Revenue'].idxmax(), 'Month']
    best_category = Revenue_by_Category.iloc[0]['Clean Category']
    worst_product = top_products.iloc[-1]['Product']

    print(f'Total Revenue: £{total_revenue:.2f}')
    print(f'Total Orders: {total_orders}')
    print(f'Average Order Value: £{avg_order_value:.2f}')
    print(f'Best Month: {best_month}')
    print(f'Best Category: {best_category}')
    print(f'Top Product: {top_product}')
    print(f'Worst Product: {worst_product}')
def build_executive_summary(df, top_product):
    total_revenue =  df['Revenue'].sum()
    total_orders = len(df)
    avg_order_value = round(total_revenue/total_orders,1)
    Executive_Summary =  pd.DataFrame({
        'Metric':[
            'Total Revenue',
            'Total Orders',
            'Average Order Value',
            'Top Product',
        ],
        'Value' : [
            total_revenue,
            total_orders,
            avg_order_value,
            top_product,
        ]
    })
    return Executive_Summary
# %%
#Data Quality Summary

def build_data_quality_summary(df,original_rows,duplicate_rows,missing_rows):
    final_rows = len(df)

    data_quality_summary = pd.DataFrame({
        'Metric': [
            'Original Rows',
            'Duplicate Rows Removed',
            'Total Missing Values',
            'Final Rows',
            ],
            'Value' : [
                original_rows,
                duplicate_rows,
                missing_rows,
                final_rows,
            ]
    })
    return data_quality_summary
# %%
#Export & Style

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment,Font

def header_style(df,Executive_Summary, Revenue_by_Category, Revenue_by_Day, Revenue_by_Month, top_products, data_quality_summary):
    with pd.ExcelWriter(output_file) as writer:
        Executive_Summary.to_excel(writer, sheet_name='Executive_Summary', index=False)
        Revenue_by_Category.to_excel(writer, sheet_name = 'Revenue_By_Category', index = False)
        Revenue_by_Day.to_excel(writer, sheet_name = 'Revenue_by_Day', index = False)
        Revenue_by_Month.to_excel(writer,sheet_name='Revenue_by_Month', index = False)
        top_products.to_excel(writer, sheet_name='Top_Products', index=False)
        df.to_excel(writer, sheet_name='Cleaned_Data', index=False)
        data_quality_summary.to_excel(writer, sheet_name='Data_Quality', index=False)

    wb = load_workbook(output_file)

    style_sheet(wb["Executive_Summary"])
    style_sheet(wb["Revenue_By_Category"])
    style_sheet(wb["Revenue_by_Day"])
    style_sheet(wb["Revenue_by_Month"])
    style_sheet(wb["Top_Products"])
    style_sheet(wb["Cleaned_Data"])
    style_sheet(wb["Data_Quality"])

    wb.save(output_file)


def style_sheet(ws):


         for cell in ws[1]:
             cell.font = Font(bold=True,color='FFFFFF')
             cell.fill = PatternFill('solid', start_color='1F3864')
             cell.alignment = Alignment(horizontal='center')

         for col in ws.columns:
             max_len = max(len(str(cell.value or ''))for cell in col)
             ws.column_dimensions[col[0].column_letter].width = max_len + 4



def int_format(ws,column_name):

    headers ={}
    for cell in ws[1]:
        headers[cell.value] = cell.column_letter

    if column_name in headers:
        for cell in ws[headers[column_name]][1:]:
            cell.number_format = '£#,##0.00'

def date_format(ws):

    headers ={}
    for cell in ws[1]:
        headers[cell.value] = cell.column_letter

    for cell in ws[headers['Clean Date']][1:]:
        cell.number_format = 'dd/mm/yyyy'


def cond_format():
            def cond_fill(ws):
                headers ={}
                for cell in ws[1]:
                    headers[cell.value]=cell.column_letter

                for row in range(2, ws.max_row + 1):
                    if ws[f"{headers['Product']}{row}"].value == top_product:

                         for cell in ws[row]:
                             cell.fill = PatternFill(
                                'solid',
                                start_color='EFBF04'
                          )

            wb = load_workbook(output_file)
            cond_fill(wb['Cleaned_Data'])
            wb.save(output_file)

def specific_format(ws):
    headers ={}
    for cell in ws[1]:
        headers[cell.value]=cell.column_letter

    for row in range(2,ws.max_row+1):
        metric=ws[f'{headers['Metric']}{row}'].value
        value_cell= ws[f'{headers['Value']}{row}']

        if metric =='Total Revenue':
            value_cell.number_format = '£#,##0.00'
        elif metric =='Average Order Value':
            value_cell.number_format = '£#,##0.00'
        elif metric =='Total Orders':
            value_cell.number_format = '0'






# %%
#Refactor

if __name__ =='__main__':
    df =  pd.read_excel(input_file)
    original_rows = len(df)
    duplicate_rows = df.duplicated().sum()
    missing_rows = df.isna().sum().sum()
    df = clean_data(df)

    Category_Summary,Revenue_by_Category,Revenue_by_Month,Revenue_by_Day, top_products,top_product = build_summaries(df)
    data_quality_summary = build_data_quality_summary(df,original_rows,duplicate_rows,missing_rows)
    Executive_Summary = build_executive_summary(df,top_product)
    header_style(df, Executive_Summary, Revenue_by_Category, Revenue_by_Day,Revenue_by_Month, top_products, data_quality_summary)
    cond_format()




    wb = load_workbook(output_file)

    int_format(wb['Cleaned_Data'], 'Revenue')
    int_format(wb['Cleaned_Data'], 'Unit_Price')
    int_format(wb['Revenue_By_Category'], 'Revenue')
    int_format(wb['Top_Products'], 'Revenue')
    int_format(wb['Executive_Summary'], 'Value')
    int_format(wb['Revenue_by_Day'], 'Revenue')
    int_format(wb['Revenue_by_Month'], 'Revenue')
    specific_format(wb['Executive_Summary'])

    date_format(wb['Cleaned_Data'])

    wb.save(output_file)

    print_summary(df,Revenue_by_Month, Revenue_by_Category,top_product)

    print('Report generated successfully!')
    print(f'📁 Output saved to: {output_file}')